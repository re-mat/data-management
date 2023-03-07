#!/usr/bin/env python

import hashlib
import logging
import os
import requests
import certifi
import json

from openpyxl.worksheet.worksheet import Worksheet
from pyclowder.extractors import Extractor
from pyclowder.utils import CheckMessage
import pyclowder.files
from openpyxl import load_workbook


class ExperimentFromExcel(Extractor):
    def __init__(self):
        Extractor.__init__(self)

        hashes = os.getenv('EXTRACTOR_HASHLIST', "md5,sha1,sha224,sha256,sha384,sha512")

        # add any additional arguments to parser
        self.parser.add_argument('--hashes', dest="hash_list", type=str, nargs='?', default=hashes,
                                 help="list of hash types to calculate")

        # parse command line and load default logging configuration
        self.setup()

        # setup logging for the exctractor
        logging.getLogger('pyclowder').setLevel(logging.DEBUG)
        logging.getLogger('__main__').setLevel(logging.DEBUG)

        # assign other arguments
        self.hash_list = [x.strip() for x in self.args.hash_list.split(',')]

    # Check whether dataset already has metadata
    def check_message(self, connector, host, secret_key, resource, parameters):
        return CheckMessage.download

    def excel_to_json(self, path):
        wb = load_workbook(filename=path)
        inputs = {}

        for sheet in wb.sheetnames:
            inputs[sheet] = []
            ws = wb[sheet]
            headers = [col.value for col in list(ws.rows)[0]]
            for row in ws.iter_rows(min_row=2):
                input_properties = {key: cell.value for key, cell in zip(headers, row)}
                inputs[sheet].append(input_properties)

        return {"inputs": inputs}

    def process_message(self, connector, host, secret_key, resource, parameters):
        logger = logging.getLogger('__main__')
        logger.info("I got a resource "+str(resource))
        experiment = self.excel_to_json(resource['local_paths'][0])
        logger.info(experiment)

        # store results as metadata
        metadata = {
            "@context": ["https://clowder.ncsa.illinois.edu/contexts/metadata.jsonld"],
            "dataset_id": resource['parent'].get('id', None),
            "content": experiment,
            "agent": {
                "@type": "cat:extractor",
                "extractor_id": host + "api/extractors/" + self.extractor_info['name']
            }
        }

        pyclowder.datasets.upload_metadata(connector, host, secret_key, resource['parent'].get('id', None), metadata)


if __name__ == "__main__":
    extractor = ExperimentFromExcel()
    # print(extractor.excel_to_json("/Users/bengal1/dev/MDF/data-management/sample.xlsx"))
    extractor.start()
